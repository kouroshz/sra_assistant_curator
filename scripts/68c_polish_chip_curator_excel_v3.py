#!/usr/bin/env python3
"""
Polish ChIP curator workbook to be closer to RNA curator-facing style.

Input:
  latest workbook pointed to by:
    outputs/06_CHIP_AI_ASSIST/21_curator_excel/LATEST_CHIP_CURATOR_REVIEW.txt

Output:
  outputs/06_CHIP_AI_ASSIST/21_curator_excel/chip_curator_review_v3_<timestamp>.xlsx

No API required.
No AI rerun.
No structured Excel tables.
"""

from pathlib import Path
from datetime import datetime
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter


OUTDIR = Path("outputs/06_CHIP_AI_ASSIST/21_curator_excel")
LATEST = OUTDIR / "LATEST_CHIP_CURATOR_REVIEW.txt"


def clean(x):
    if pd.isna(x):
        return ""
    s = str(x).strip()
    if s.lower() in {"nan", "none", "<na>"}:
        return ""
    return s


def load_latest_workbook_path():
    if not LATEST.exists():
        raise SystemExit(f"Missing latest pointer: {LATEST}")
    p = Path(LATEST.read_text().strip())
    if not p.exists():
        raise SystemExit(f"Latest workbook does not exist: {p}")
    return p


def action_hint(problem_type):
    p = clean(problem_type).lower()

    if p == "low_or_missing_confidence":
        return "Review AI call carefully; use source metadata/evidence to confirm or correct."
    if p == "review_flag_curator_check":
        return "Curator should inspect this row because AI/deterministic repair flagged it."
    if p == "unknown_or_missing_sample_role":
        return "Assign final sample role: target_ip, input, IgG, untagged_control, mock, or control_sample."
    if p == "target_ip_missing_background":
        return "Verify whether this target-IP row has a matching input/IgG/background control."
    if p == "packet_peak_calling_readiness":
        return "Review study-level ChIP readiness and whether missing/partial controls block peak calling."
    if p == "unknown_condition":
        return "Fill final condition if recoverable from paper/sample metadata; otherwise mark as unknown/NA."
    if p == "unknown_stage":
        return "Fill final stage/timepoint if recoverable from paper/sample metadata; otherwise mark as unknown/NA."
    if "missing_background" in p:
        return "Resolve target-control relationship or mark as not peak-calling-ready."
    if "control_run_not_resolved" in p:
        return "Check assigned control SRR and whether it exists as a row in this packet or external metadata."
    if "check_control_role" in p:
        return "Verify that the resolved control row is truly input/IgG/background, not target-IP."

    return "Review source metadata, AI evidence, and curator notes before finalizing."


def problem_message(row):
    p = clean(row.get("problem_type", ""))
    role = clean(row.get("suggested_sample_role", ""))
    target = clean(row.get("suggested_target_or_antibody_or_tag", "")) or clean(row.get("target_clean", ""))
    run = clean(row.get("Run", ""))
    conf = clean(row.get("suggestion_confidence", ""))
    flag = clean(row.get("review_flag", ""))
    bg = clean(row.get("suggested_comparator_or_background", ""))
    prelim_bg = clean(row.get("matched_background_run_ids_prelim", ""))

    pieces = [f"{p}"]
    if run:
        pieces.append(f"Run={run}")
    if target:
        pieces.append(f"target={target}")
    if role:
        pieces.append(f"role={role}")
    if conf:
        pieces.append(f"confidence={conf}")
    if flag:
        pieces.append(f"flag={flag}")
    if bg or prelim_bg:
        pieces.append(f"AI_bg={bg or 'blank'}; prelim_bg={prelim_bg or 'blank'}")

    return "; ".join(pieces)


def study_curator_summary(row):
    targets = clean(row.get("targets", ""))
    peak = clean(row.get("chip_peak_calling_ready", ""))
    low = clean(row.get("n_low_confidence_rows", ""))
    review = clean(row.get("n_review_flag_rows", ""))
    missing_bg = clean(row.get("n_target_ip_missing_background", ""))
    unknown_cond = clean(row.get("n_unknown_condition_rows", ""))

    parts = []
    if targets:
        parts.append(f"Targets: {targets[:180]}")
    if peak:
        parts.append(f"Peak-calling readiness: {peak}")
    if low and low != "0":
        parts.append(f"low-confidence rows={low}")
    if review and review != "0":
        parts.append(f"review-flag rows={review}")
    if missing_bg and missing_bg != "0":
        parts.append(f"target-IP missing background={missing_bg}")
    if unknown_cond and unknown_cond != "0":
        parts.append(f"unknown condition rows={unknown_cond}")

    return " | ".join(parts)


def study_curator_priority(row):
    peak = clean(row.get("chip_peak_calling_ready", "")).lower()
    low = int(float(clean(row.get("n_low_confidence_rows", "0")) or 0))
    review = int(float(clean(row.get("n_review_flag_rows", "0")) or 0))
    missing_bg = int(float(clean(row.get("n_target_ip_missing_background", "0")) or 0))

    if peak == "no" or missing_bg > 0 or low > 0:
        return "HIGH_REVIEW"
    if peak == "partial" or review > 0:
        return "REVIEW"
    return "OK"


def add_or_update_columns(df, sheet_name):
    df = df.copy()

    if sheet_name == "Study_Review":
        df.insert(0, "curator_priority", df.apply(study_curator_priority, axis=1))
        df.insert(1, "curator_focus_summary", df.apply(study_curator_summary, axis=1))

    if sheet_name in {"Problem_Rows", "Metadata_Gaps"}:
        if "problem_message" not in df.columns:
            df.insert(2, "problem_message", df.apply(problem_message, axis=1))
        if "curator_action_hint" not in df.columns:
            df.insert(3, "curator_action_hint", df["problem_type"].map(action_hint) if "problem_type" in df.columns else "")

    return df


def safe_sheet_name(name):
    return name[:31]


def write_v3_workbook(input_xlsx):
    sheets = pd.read_excel(input_xlsx, sheet_name=None, dtype=str)
    sheets = {name: df.fillna("") for name, df in sheets.items()}

    polished = {}
    for name, df in sheets.items():
        polished[name] = add_or_update_columns(df, name)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = OUTDIR / f"chip_curator_review_v3_{ts}.xlsx"

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        for name, df in polished.items():
            df.to_excel(writer, sheet_name=safe_sheet_name(name), index=False)

    return out, polished


def style_workbook(path):
    wb = load_workbook(path)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    thin_gray = Side(style="thin", color="D9D9D9")
    border = Border(bottom=thin_gray)

    wrap_keywords = [
        "summary", "message", "hint", "evidence", "metadata", "warning",
        "notes", "comment", "json", "source_row_ids", "run_ids",
        "blockers", "description", "targets", "matched", "resolved"
    ]

    freeze_map = {
        "README": "A2",
        "QC_Summary": "A2",
        "Study_Review": "D2",
        "Target_Control_Map_Review": "G2",
        "Problem_Rows": "H2",
        "Metadata_Gaps": "H2",
        "Rowwise_Review": "G2",
        "Sample_Map_Review": "E2",
        "Technical_Inventory": "D2",
    }

    for ws in wb.worksheets:
        ws.freeze_panes = freeze_map.get(ws.title, "A2")

        if ws.max_row >= 2 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions

        # Header styling
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[1].height = 30

        header_to_col = {str(ws.cell(row=1, column=i).value): i for i in range(1, ws.max_column + 1)}

        # Widths and wrapping
        for idx in range(1, ws.max_column + 1):
            col_name = str(ws.cell(row=1, column=idx).value)
            col_l = col_name.lower()

            sample_vals = [col_name]
            for ridx in range(2, min(ws.max_row, 200) + 1):
                sample_vals.append(str(ws.cell(row=ridx, column=idx).value or ""))

            max_len = max(len(v) for v in sample_vals)
            should_wrap = any(k in col_l for k in wrap_keywords)

            if should_wrap:
                width = min(max(max_len + 2, 22), 60)
            else:
                width = min(max(max_len + 2, 10), 32)

            ws.column_dimensions[get_column_letter(idx)].width = width

            for cell in ws[get_column_letter(idx)]:
                cell.alignment = Alignment(vertical="top", wrap_text=should_wrap)

        for ridx in range(2, ws.max_row + 1):
            ws.row_dimensions[ridx].height = 18

        def col_letter(col_name):
            idx = header_to_col.get(col_name)
            return get_column_letter(idx) if idx else None

        def add_contains(col_name, text, fill, font_color="000000"):
            col = col_letter(col_name)
            if not col or ws.max_row < 2:
                return
            rng = f"{col}2:{col}{ws.max_row}"
            formula = f'IFERROR(ISNUMBER(SEARCH("{text}",{col}2)),FALSE)'
            ws.conditional_formatting.add(
                rng,
                FormulaRule(
                    formula=[formula],
                    fill=PatternFill("solid", fgColor=fill),
                    font=Font(color=font_color)
                )
            )

        # Confidence
        for c in [
            "suggestion_confidence", "confidence", "target_confidence",
            "resolved_control_confidence"
        ]:
            add_contains(c, "low", "F4CCCC")
            add_contains(c, "medium", "FFF2CC")
            add_contains(c, "high", "D9EAD3")

        # Status / review
        for c in [
            "curator_priority", "severity", "review_flag", "target_review_flag",
            "resolved_control_review_flag", "warning_flags",
            "target_control_match_status", "validation_status",
            "chip_peak_calling_ready", "analysis_ready_status"
        ]:
            add_contains(c, "HIGH_REVIEW", "F4CCCC")
            add_contains(c, "FAIL", "F4CCCC")
            add_contains(c, "missing", "F4CCCC")
            add_contains(c, "low", "F4CCCC")
            add_contains(c, "no", "F4CCCC")
            add_contains(c, "REVIEW", "FFF2CC")
            add_contains(c, "partial", "FFF2CC")
            add_contains(c, "curator_check", "FFF2CC")
            add_contains(c, "ambiguous", "FCE5CD")
            add_contains(c, "PASS", "D9EAD3")
            add_contains(c, "OK", "D9EAD3")
            add_contains(c, "ok", "D9EAD3")
            add_contains(c, "yes", "D9EAD3")

    wb.save(path)


def validate_no_tables(path):
    import zipfile
    with zipfile.ZipFile(path) as z:
        tables = [n for n in z.namelist() if n.startswith("xl/tables/")]
        cf_count = 0
        for n in z.namelist():
            if n.startswith("xl/worksheets/") and n.endswith(".xml"):
                txt = z.read(n).decode("utf-8", errors="ignore")
                cf_count += txt.count("<conditionalFormatting")
        return len(tables), cf_count


def main():
    input_xlsx = load_latest_workbook_path()
    out, polished = write_v3_workbook(input_xlsx)
    style_workbook(out)

    tables, cf_count = validate_no_tables(out)

    LATEST.write_text(str(out) + "\n")

    summary = OUTDIR / f"{out.stem}.summary.txt"
    lines = []
    lines.append("ChIP curator workbook V3")
    lines.append(f"Generated: {datetime.now().isoformat(timespec='seconds')}")
    lines.append(f"Input: {input_xlsx}")
    lines.append(f"Workbook: {out}")
    lines.append("")
    for name, df in polished.items():
        lines.append(f"{name}: {len(df)} rows, {len(df.columns)} columns")
    lines.append("")
    lines.append(f"Excel table parts: {tables}")
    lines.append(f"Conditional formatting blocks: {cf_count}")
    lines.append("")
    lines.append("V3 improvements:")
    lines.append("- RNA-style frozen identifier columns.")
    lines.append("- Confidence/status/review highlighting.")
    lines.append("- Study_Review has curator_priority and curator_focus_summary.")
    lines.append("- Problem_Rows and Metadata_Gaps have problem_message and curator_action_hint.")
    lines.append("- No structured Excel tables, avoiding Excel repair warnings.")
    summary.write_text("\n".join(lines))

    print("Wrote:", out)
    print("Wrote:", LATEST)
    print("Wrote:", summary)
    print()
    print("\n".join(lines))


if __name__ == "__main__":
    main()
