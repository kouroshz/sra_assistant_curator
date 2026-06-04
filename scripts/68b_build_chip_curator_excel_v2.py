#!/usr/bin/env python3
"""
Build a more curator-facing ChIP review workbook.

This improves the first ChIP workbook by mirroring the RNA curator-facing logic:

- richer Study_Review
- richer Rowwise_Review with source metadata
- Target_Control_Map_Review includes confidence, review flags, evidence, and resolved control rows
- detailed Problem_Rows
- structured curator_final_* columns
- technical paths moved to Technical_Inventory instead of primary review sheets

No API required.
"""

from pathlib import Path
from datetime import datetime
from collections import Counter, defaultdict
import json
import re
import pandas as pd

OUTDIR = Path("outputs/06_CHIP_AI_ASSIST/21_curator_excel")
OUTDIR.mkdir(parents=True, exist_ok=True)

ACTIVE = Path("outputs/06_CHIP_AI_ASSIST/14_chip_ai_inventory/chip_ai_active_validated_outputs.tsv")
FINAL = Path("outputs/06_CHIP_AI_ASSIST/20_final_qc")


def clean(x):
    if pd.isna(x):
        return ""
    s = str(x).strip()
    if s.lower() in {"nan", "none", "<na>"}:
        return ""
    return s


def norm(x):
    return clean(x).lower().replace("-", "_").replace(" ", "_")


def split_runs(x):
    x = clean(x)
    if not x:
        return []
    parts = re.split(r"[;,|]\s*|\s+", x)
    return [p.strip() for p in parts if p.strip() and p.strip().lower() not in {"na", "none", "unknown"}]


def short_json(x, max_len=1200):
    if isinstance(x, (dict, list)):
        return json.dumps(x, ensure_ascii=False)[:max_len]
    return clean(x)[:max_len]


def read_json(path):
    try:
        return json.loads(Path(path).read_text())
    except Exception:
        return {}


def classify_problem(row):
    problems = []

    role = norm(row.get("ai_sample_role", ""))
    conf = norm(row.get("suggestion_confidence", row.get("confidence", "")))
    flag = norm(row.get("review_flag", ""))
    bg = norm(row.get("ai_background_or_comparator", ""))
    prelim_bg = clean(row.get("matched_background_run_ids_prelim", ""))

    if conf in {"low", "unknown", ""}:
        problems.append("low_or_missing_confidence")

    if flag not in {"", "ok"}:
        problems.append(f"review_flag_{flag}")

    if role in {"", "unknown"}:
        problems.append("unknown_or_missing_sample_role")

    if role == "target_ip" and bg in {"", "unknown", "not_applicable", "na", "n_a"} and prelim_bg == "":
        problems.append("target_ip_missing_background")

    if clean(row.get("condition", "")).lower() in {"unknown", ""}:
        problems.append("unknown_condition")

    if clean(row.get("stage", "")).lower() in {"unknown", ""}:
        problems.append("unknown_stage")

    return problems


def build_readme():
    return pd.DataFrame([
        {
            "section": "Purpose",
            "description": "This workbook contains AI-assisted ChIP/CUT&RUN/CUT&Tag metadata suggestions for curator review. AI suggestions are not final; curator_final_* columns are authoritative."
        },
        {
            "section": "Recommended workflow",
            "description": "Start with Study_Review. Then inspect Problem_Rows. For ChIP, prioritize Target_Control_Map_Review because target-IP to input/IgG/background mapping is central. Then spot-check Rowwise_Review and Sample_Map_Review."
        },
        {
            "section": "Confidence",
            "description": "confidence/suggestion_confidence comes from AI or deterministic repair. Low and medium confidence rows should be reviewed. Rows with review_flag != ok are surfaced in Problem_Rows."
        },
        {
            "section": "ChIP controls",
            "description": "Input/background/IgG rows are usually separate SRR rows. Shared input controls are expected and should be represented in Target_Control_Map_Review, not by duplicate sample_map membership."
        },
        {
            "section": "Curator columns",
            "description": "Use curator_final_* and curator_notes/status columns for corrections. Do not overwrite source metadata or AI suggestion/evidence columns."
        },
    ])


def build_all_tables():
    active = pd.read_csv(ACTIVE, sep="\t", dtype=str).fillna("")

    study_rows = []
    rowwise_rows = []
    sample_map_rows = []
    target_control_rows = []
    problem_rows = []

    for _, inv in active.iterrows():
        packet_id = clean(inv.get("packet_id", ""))
        ai_json = clean(inv.get("active_ai_json", ""))
        packet_table_path = clean(inv.get("packet_table", ""))

        obj = read_json(ai_json)
        table = pd.read_csv(packet_table_path, sep="\t", dtype=str).fillna("") if packet_table_path and Path(packet_table_path).exists() else pd.DataFrame()

        table_by_sid = {}
        table_by_run = defaultdict(list)
        if not table.empty and "source_row_id" in table.columns:
            for _, tr in table.iterrows():
                sid = clean(tr.get("source_row_id", ""))
                run = clean(tr.get("Run", ""))
                table_by_sid[sid] = tr.to_dict()
                if run:
                    table_by_run[run].append(tr.to_dict())

        rowwise = obj.get("rowwise_suggestions", []) or []
        rowwise_by_sid = {clean(r.get("source_row_id", "")): r for r in rowwise}

        role_counts = Counter()
        confidence_counts = Counter()
        flag_counts = Counter()
        n_low = 0
        n_review = 0
        n_unknown_condition = 0
        n_target_ip_missing_bg = 0

        for rw in rowwise:
            sid = clean(rw.get("source_row_id", ""))
            src = table_by_sid.get(sid, {})

            role = clean(rw.get("suggested_sample_role", ""))
            conf = clean(rw.get("suggestion_confidence", ""))
            flag = clean(rw.get("review_flag", ""))
            stage = clean(rw.get("suggested_stage_timepoint", ""))
            condition = clean(rw.get("suggested_condition", ""))
            bg = clean(rw.get("suggested_comparator_or_background", ""))

            role_counts[role or "blank"] += 1
            confidence_counts[conf or "blank"] += 1
            flag_counts[flag or "blank"] += 1

            if norm(conf) in {"low", "unknown", ""}:
                n_low += 1
            if norm(flag) not in {"", "ok"}:
                n_review += 1
            if norm(condition) in {"", "unknown"}:
                n_unknown_condition += 1

            prelim_bg = clean(src.get("matched_background_run_ids_prelim", ""))
            if norm(role) == "target_ip" and norm(bg) in {"", "unknown", "not_applicable"} and prelim_bg == "":
                n_target_ip_missing_bg += 1

            row = {
                "packet_id": packet_id,
                "pmid": clean(inv.get("pmid", "")),
                "bioproject": clean(inv.get("bioproject", "")),
                "source_row_id": sid,
                "Run": clean(src.get("Run", rw.get("Run", ""))),
                "BioSample": clean(src.get("BioSample", "")),
                "original_Target": clean(src.get("Target", "")),
                "target_clean": clean(src.get("target_clean", "")),
                "target_type": clean(src.get("target_type", "")),
                "chip_role_original": clean(src.get("chip_role_original", "")),
                "chip_role_for_ai": clean(src.get("chip_role_for_ai", "")),
                "sample_role_prelim": clean(src.get("sample_role_prelim", "")),
                "control_type_prelim": clean(src.get("control_type_prelim", "")),
                "background_sample": clean(src.get("background_sample", "")),
                "assigned_control1": clean(src.get("assigned_control1", "")),
                "assigned_control2": clean(src.get("assigned_control2", "")),
                "matched_background_run_ids_prelim": prelim_bg,
                "raw_metadata_joined": clean(src.get("raw_metadata_joined", "")),
                "public_metadata_evidence_compact": clean(src.get("public_metadata_evidence_compact", "")),
                "sample_class_id": clean(rw.get("sample_class_id", "")),
                "suggested_sample_role": role,
                "suggested_target_or_antibody_or_tag": clean(rw.get("suggested_target_or_antibody_or_tag", "")),
                "suggested_comparator_or_background": bg,
                "suggested_stage_timepoint": stage,
                "suggested_strain": clean(rw.get("suggested_strain", "")),
                "suggested_condition": condition,
                "suggestion_confidence": conf,
                "review_flag": flag,
                "suggestion_evidence": clean(rw.get("suggestion_evidence", "")),
                "curator_final_target_or_antibody": "",
                "curator_final_sample_role": "",
                "curator_final_stage": "",
                "curator_final_strain": "",
                "curator_final_condition": "",
                "curator_final_control_or_background": "",
                "curator_row_status": "",
                "curator_notes": "",
            }
            rowwise_rows.append(row)

            problems = classify_problem({
                **row,
                "confidence": conf,
                "ai_sample_role": role,
                "ai_background_or_comparator": bg,
                "stage": stage,
                "condition": condition,
            })

            for p in problems:
                severity = "REVIEW"
                if p in {"low_or_missing_confidence", "unknown_or_missing_sample_role", "target_ip_missing_background"}:
                    severity = "HIGH_REVIEW"

                problem_rows.append({
                    "problem_type": p,
                    "severity": severity,
                    "packet_id": packet_id,
                    "pmid": clean(inv.get("pmid", "")),
                    "bioproject": clean(inv.get("bioproject", "")),
                    "source_row_id": sid,
                    "Run": row["Run"],
                    "BioSample": row["BioSample"],
                    "target_clean": row["target_clean"],
                    "suggested_target_or_antibody_or_tag": row["suggested_target_or_antibody_or_tag"],
                    "sample_role_prelim": row["sample_role_prelim"],
                    "suggested_sample_role": role,
                    "stage": stage,
                    "strain": row["suggested_strain"],
                    "condition": condition,
                    "suggested_comparator_or_background": bg,
                    "matched_background_run_ids_prelim": prelim_bg,
                    "assigned_control1": row["assigned_control1"],
                    "assigned_control2": row["assigned_control2"],
                    "suggestion_confidence": conf,
                    "review_flag": flag,
                    "suggestion_evidence": row["suggestion_evidence"],
                    "raw_metadata_joined": row["raw_metadata_joined"],
                    "curator_problem_status": "",
                    "curator_corrected_value": "",
                    "curator_notes": "",
                })

        # Sample map rows
        for sm in obj.get("sample_map", []) or []:
            sample_map_rows.append({
                "packet_id": packet_id,
                "pmid": clean(inv.get("pmid", "")),
                "bioproject": clean(inv.get("bioproject", "")),
                "sample_class_id": clean(sm.get("sample_class_id", "")),
                "sample_class_description": clean(sm.get("sample_class_description", "")),
                "sample_role": clean(sm.get("sample_role", "")),
                "target_or_antibody_or_tag": clean(sm.get("target_or_antibody_or_tag", "")),
                "stage_or_timepoint": clean(sm.get("stage_or_timepoint", "")),
                "strain": clean(sm.get("strain", "")),
                "condition": clean(sm.get("condition", "")),
                "perturbation_or_treatment": clean(sm.get("perturbation_or_treatment", "")),
                "n_rows_matched": clean(sm.get("n_rows_matched", "")),
                "matched_source_row_ids": short_json(sm.get("matched_source_row_ids", "")),
                "matched_run_ids": short_json(sm.get("matched_run_ids", "")),
                "suggested_comparator_or_background_class_id": clean(sm.get("suggested_comparator_or_background_class_id", "")),
                "analysis_ready_status": clean(sm.get("analysis_ready_status", "")),
                "blocker_reason": clean(sm.get("blocker_reason", "")),
                "confidence": clean(sm.get("confidence", "")),
                "curator_check_priority": clean(sm.get("curator_check_priority", "")),
                "warning_flags": short_json(sm.get("warning_flags", "")),
                "evidence": clean(sm.get("evidence", "")),
                "curator_final_sample_class_id": "",
                "curator_final_target_or_antibody": "",
                "curator_final_sample_role": "",
                "curator_final_stage": "",
                "curator_final_strain": "",
                "curator_final_condition": "",
                "curator_final_control_or_background": "",
                "curator_sample_status": "",
                "curator_notes": "",
            })

        # Target-control map, resolved to control row(s)
        for rw in rowwise:
            sid = clean(rw.get("source_row_id", ""))
            role = norm(rw.get("suggested_sample_role", ""))
            if role != "target_ip":
                continue

            src = table_by_sid.get(sid, {})
            prelim_runs = split_runs(src.get("matched_background_run_ids_prelim", "")) or split_runs(src.get("assigned_control1", "")) + split_runs(src.get("assigned_control2", ""))

            control_source_ids = []
            control_roles = []
            control_classes = []
            control_stages = []
            control_strains = []
            control_conditions = []
            control_conf = []
            control_flags = []
            control_evidence = []

            for crun in prelim_runs:
                for csrc in table_by_run.get(crun, []):
                    csid = clean(csrc.get("source_row_id", ""))
                    crw = rowwise_by_sid.get(csid, {})
                    control_source_ids.append(csid)
                    control_roles.append(clean(crw.get("suggested_sample_role", csrc.get("sample_role_prelim", ""))))
                    control_classes.append(clean(crw.get("sample_class_id", "")))
                    control_stages.append(clean(crw.get("suggested_stage_timepoint", csrc.get("stage_combined", ""))))
                    control_strains.append(clean(crw.get("suggested_strain", csrc.get("strain_context", ""))))
                    control_conditions.append(clean(crw.get("suggested_condition", csrc.get("condition_context", ""))))
                    control_conf.append(clean(crw.get("suggestion_confidence", "")))
                    control_flags.append(clean(crw.get("review_flag", "")))
                    control_evidence.append(clean(crw.get("suggestion_evidence", "")))

            ai_bg = clean(rw.get("suggested_comparator_or_background", ""))
            match_status = "ok"
            if not prelim_runs and norm(ai_bg) in {"", "unknown", "not_applicable"}:
                match_status = "missing_background"
            elif not control_source_ids and prelim_runs:
                match_status = "control_run_not_resolved_to_row"
            elif any(norm(x) not in {"input", "igg", "control_sample", "untagged_control", "mock"} for x in control_roles if x):
                match_status = "check_control_role"

            target_control_rows.append({
                "packet_id": packet_id,
                "pmid": clean(inv.get("pmid", "")),
                "bioproject": clean(inv.get("bioproject", "")),
                "target_source_row_id": sid,
                "target_Run": clean(src.get("Run", rw.get("Run", ""))),
                "target_BioSample": clean(src.get("BioSample", "")),
                "original_Target": clean(src.get("Target", "")),
                "target_clean": clean(src.get("target_clean", "")),
                "suggested_target_or_antibody_or_tag": clean(rw.get("suggested_target_or_antibody_or_tag", "")),
                "target_sample_class_id": clean(rw.get("sample_class_id", "")),
                "target_stage": clean(rw.get("suggested_stage_timepoint", "")),
                "target_strain": clean(rw.get("suggested_strain", "")),
                "target_condition": clean(rw.get("suggested_condition", "")),
                "target_confidence": clean(rw.get("suggestion_confidence", "")),
                "target_review_flag": clean(rw.get("review_flag", "")),
                "target_evidence": clean(rw.get("suggestion_evidence", "")),
                "ai_background_or_comparator": ai_bg,
                "prelim_matched_background_run_ids": ";".join(prelim_runs),
                "assigned_control1": clean(src.get("assigned_control1", "")),
                "assigned_control2": clean(src.get("assigned_control2", "")),
                "resolved_control_source_row_ids": ";".join(sorted(set(control_source_ids))),
                "resolved_control_roles": ";".join(sorted(set(x for x in control_roles if x))),
                "resolved_control_sample_class_ids": ";".join(sorted(set(x for x in control_classes if x))),
                "resolved_control_stage": ";".join(sorted(set(x for x in control_stages if x))),
                "resolved_control_strain": ";".join(sorted(set(x for x in control_strains if x))),
                "resolved_control_condition": ";".join(sorted(set(x for x in control_conditions if x))),
                "resolved_control_confidence": ";".join(sorted(set(x for x in control_conf if x))),
                "resolved_control_review_flag": ";".join(sorted(set(x for x in control_flags if x))),
                "resolved_control_evidence": short_json(" | ".join(x for x in control_evidence if x)),
                "target_control_match_status": match_status,
                "curator_final_background_run_ids": "",
                "curator_final_background_role": "",
                "curator_target_control_status": "",
                "curator_notes": "",
            })

            if match_status != "ok":
                problem_rows.append({
                    "problem_type": match_status,
                    "severity": "HIGH_REVIEW",
                    "packet_id": packet_id,
                    "pmid": clean(inv.get("pmid", "")),
                    "bioproject": clean(inv.get("bioproject", "")),
                    "source_row_id": sid,
                    "Run": clean(src.get("Run", "")),
                    "BioSample": clean(src.get("BioSample", "")),
                    "target_clean": clean(src.get("target_clean", "")),
                    "suggested_target_or_antibody_or_tag": clean(rw.get("suggested_target_or_antibody_or_tag", "")),
                    "sample_role_prelim": clean(src.get("sample_role_prelim", "")),
                    "suggested_sample_role": clean(rw.get("suggested_sample_role", "")),
                    "stage": clean(rw.get("suggested_stage_timepoint", "")),
                    "strain": clean(rw.get("suggested_strain", "")),
                    "condition": clean(rw.get("suggested_condition", "")),
                    "suggested_comparator_or_background": ai_bg,
                    "matched_background_run_ids_prelim": ";".join(prelim_runs),
                    "assigned_control1": clean(src.get("assigned_control1", "")),
                    "assigned_control2": clean(src.get("assigned_control2", "")),
                    "suggestion_confidence": clean(rw.get("suggestion_confidence", "")),
                    "review_flag": clean(rw.get("review_flag", "")),
                    "suggestion_evidence": clean(rw.get("suggestion_evidence", "")),
                    "raw_metadata_joined": clean(src.get("raw_metadata_joined", "")),
                    "curator_problem_status": "",
                    "curator_corrected_value": "",
                    "curator_notes": "",
                })

        ar = obj.get("analysis_readiness", {}) or {}
        peak = clean(inv.get("chip_peak_calling_ready", ar.get("chip_peak_calling_ready", "")))

        study_rows.append({
            "packet_id": packet_id,
            "pmid": clean(inv.get("pmid", "")),
            "bioproject": clean(inv.get("bioproject", "")),
            "assay_class_confirmed": clean(obj.get("assay_class_confirmed", "")),
            "n_rows": clean(inv.get("n_rows", "")),
            "targets": clean(inv.get("targets", "")),
            "target_types": clean(inv.get("target_types", "")),
            "validation_status": clean(inv.get("validation_status", "")),
            "chip_peak_calling_ready": peak,
            "n_rowwise_suggestions": len(rowwise),
            "n_sample_map_entries": len(obj.get("sample_map", []) or []),
            "n_target_ip_rows": role_counts.get("target_ip", 0),
            "n_input_rows": role_counts.get("input", 0),
            "n_IgG_rows": role_counts.get("IgG", 0) + role_counts.get("igg", 0),
            "n_low_confidence_rows": n_low,
            "n_review_flag_rows": n_review,
            "n_unknown_condition_rows": n_unknown_condition,
            "n_target_ip_missing_background": n_target_ip_missing_bg,
            "confidence_counts": short_json(dict(confidence_counts)),
            "review_flag_counts": short_json(dict(flag_counts)),
            "main_blockers": short_json(ar.get("main_blockers", "")),
            "global_warnings": short_json(obj.get("global_warnings", "")),
            "curator_study_status": "",
            "curator_ready_for_row_review": "",
            "curator_study_notes": "",
            "technical_active_ai_json": ai_json,
        })

        if peak in {"partial", "no", "unknown", ""}:
            problem_rows.append({
                "problem_type": "packet_peak_calling_readiness",
                "severity": "REVIEW",
                "packet_id": packet_id,
                "pmid": clean(inv.get("pmid", "")),
                "bioproject": clean(inv.get("bioproject", "")),
                "source_row_id": "",
                "Run": "",
                "BioSample": "",
                "target_clean": clean(inv.get("targets", "")),
                "suggested_target_or_antibody_or_tag": "",
                "sample_role_prelim": "",
                "suggested_sample_role": "",
                "stage": "",
                "strain": "",
                "condition": "",
                "suggested_comparator_or_background": "",
                "matched_background_run_ids_prelim": "",
                "assigned_control1": "",
                "assigned_control2": "",
                "suggestion_confidence": "",
                "review_flag": "",
                "suggestion_evidence": f"chip_peak_calling_ready={peak}",
                "raw_metadata_joined": "",
                "curator_problem_status": "",
                "curator_corrected_value": "",
                "curator_notes": "",
            })

    study = pd.DataFrame(study_rows)
    sample_map = pd.DataFrame(sample_map_rows)
    target_control = pd.DataFrame(target_control_rows)
    rowwise_df = pd.DataFrame(rowwise_rows)
    problems = pd.DataFrame(problem_rows)

    # Put most important problems first.
    if not problems.empty:
        sev_order = {"HIGH_REVIEW": 0, "REVIEW": 1}
        problems["_sev_order"] = problems["severity"].map(sev_order).fillna(9)
        problems = problems.sort_values(["_sev_order", "problem_type", "packet_id", "Run"]).drop(columns=["_sev_order"])

    return active, study, sample_map, target_control, problems, rowwise_df


def autosize(writer, sheets):
    """
    Excel-safe formatting:
      - styled headers
      - frozen header row
      - plain worksheet autofilter
      - no structured Excel Table objects

    Reason: Excel repaired V2 because worksheet-level AutoFilter and
    structured Table AutoFilter were both present.
    """
    wb = writer.book

    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    header_fill = "1F4E78"
    header_font = "FFFFFF"

    wrap_keywords = [
        "evidence", "metadata", "warning", "notes", "comment",
        "json", "source_row_ids", "run_ids", "blockers",
        "description", "targets", "matched", "resolved"
    ]

    for sheet_name, df in sheets.items():
        ws = wb[sheet_name]
        ws.freeze_panes = "A2"

        if ws.max_row >= 2 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions

        # Header style
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor=header_fill)
            cell.font = Font(bold=True, color=header_font)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.row_dimensions[1].height = 28

        # Column widths and wrapping
        for idx, col in enumerate(df.columns, start=1):
            col_l = str(col).lower()
            values = [str(col)] + [str(x) for x in df[col].head(250).fillna("")]
            max_len = max(len(v) for v in values) if values else len(str(col))

            if any(k in col_l for k in wrap_keywords):
                width = min(max(max_len + 2, 20), 55)
                wrap = True
            else:
                width = min(max(max_len + 2, 10), 28)
                wrap = False

            ws.column_dimensions[get_column_letter(idx)].width = width

            # Avoid huge row heights from wrapping every cell.
            for cell in ws[get_column_letter(idx)]:
                cell.alignment = Alignment(vertical="top", wrap_text=wrap)

        # Reasonable row heights; do not let long evidence text dominate.
        for ridx in range(2, ws.max_row + 1):
            ws.row_dimensions[ridx].height = 18

        # RNA-style curator highlighting.
        # Keep this simple and Excel-safe: worksheet-level conditional formatting only.
        header_to_col = {str(ws.cell(row=1, column=i).value): i for i in range(1, ws.max_column + 1)}

        def col_letter(col_name):
            idx = header_to_col.get(col_name)
            if not idx:
                return None
            return get_column_letter(idx)

        def add_contains(col_name, text, fill, font_color="000000"):
            col = col_letter(col_name)
            if not col or ws.max_row < 2:
                return
            from openpyxl.formatting.rule import CellIsRule, FormulaRule
            from openpyxl.styles import PatternFill, Font

            rng = f"{col}2:{col}{ws.max_row}"
            # SEARCH is case-insensitive; IFERROR avoids errors on blanks.
            formula = f'IFERROR(ISNUMBER(SEARCH("{text}",{col}2)),FALSE)'
            ws.conditional_formatting.add(
                rng,
                FormulaRule(
                    formula=[formula],
                    fill=PatternFill("solid", fgColor=fill),
                    font=Font(color=font_color)
                )
            )

        # Confidence highlighting
        for c in ["suggestion_confidence", "confidence", "target_confidence", "resolved_control_confidence"]:
            add_contains(c, "low", "F4CCCC")       # red/pink
            add_contains(c, "medium", "FFF2CC")    # yellow
            add_contains(c, "high", "D9EAD3")      # green

        # Review flags / warnings
        for c in ["review_flag", "target_review_flag", "resolved_control_review_flag", "warning_flags"]:
            add_contains(c, "ok", "D9EAD3")
            add_contains(c, "curator_check", "FFF2CC")
            add_contains(c, "missing_background", "F4CCCC")
            add_contains(c, "ambiguous", "FCE5CD")
            add_contains(c, "low", "F4CCCC")

        # Validation / readiness / match status
        for c in ["validation_status", "chip_peak_calling_ready", "analysis_ready_status", "target_control_match_status"]:
            add_contains(c, "PASS", "D9EAD3")
            add_contains(c, "yes", "D9EAD3")
            add_contains(c, "ok", "D9EAD3")
            add_contains(c, "partial", "FFF2CC")
            add_contains(c, "no", "F4CCCC")
            add_contains(c, "missing", "F4CCCC")
            add_contains(c, "FAIL", "F4CCCC")

        # Problem severity
        for c in ["severity"]:
            add_contains(c, "HIGH_REVIEW", "F4CCCC")
            add_contains(c, "REVIEW", "FFF2CC")


def main():
    active, study, sample_map, target_control, problems, rowwise_df = build_all_tables()

    readme = build_readme()

    # Split low-priority metadata gaps out of actionable Problem_Rows.
    metadata_gap_types = {"unknown_condition", "unknown_stage"}

    if not problems.empty and "problem_type" in problems.columns:
        metadata_gaps = problems[problems["problem_type"].isin(metadata_gap_types)].copy()
        problems_actionable = problems[~problems["problem_type"].isin(metadata_gap_types)].copy()
    else:
        metadata_gaps = pd.DataFrame()
        problems_actionable = problems.copy()

    # Compact QC summary for curators.
    qc_rows = []

    def add_metric(section, metric, value):
        qc_rows.append({"section": section, "metric": metric, "value": value})

    add_metric("Workbook", "Study packets", len(study))
    add_metric("Workbook", "Target-control rows", len(target_control))
    add_metric("Workbook", "Rowwise rows", len(rowwise_df))
    add_metric("Workbook", "Sample-map rows", len(sample_map))
    add_metric("Workbook", "Actionable problem rows", len(problems_actionable))
    add_metric("Workbook", "Metadata gap rows", len(metadata_gaps))

    if "severity" in problems_actionable.columns and not problems_actionable.empty:
        for k, v in problems_actionable["severity"].value_counts().items():
            add_metric("Actionable Problem_Rows", str(k), int(v))

    if "problem_type" in problems_actionable.columns and not problems_actionable.empty:
        for k, v in problems_actionable["problem_type"].value_counts().items():
            add_metric("Actionable Problem Types", str(k), int(v))

    if "suggestion_confidence" in rowwise_df.columns:
        for k, v in rowwise_df["suggestion_confidence"].value_counts().items():
            add_metric("Rowwise confidence", str(k), int(v))

    if "review_flag" in rowwise_df.columns:
        for k, v in rowwise_df["review_flag"].value_counts().items():
            add_metric("Rowwise review flags", str(k), int(v))

    if "target_control_match_status" in target_control.columns:
        for k, v in target_control["target_control_match_status"].value_counts().items():
            add_metric("Target-control match status", str(k), int(v))

    qc_summary = pd.DataFrame(qc_rows)

    # Technical inventory is kept, but last.
    technical = active.copy()

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = OUTDIR / f"chip_curator_review_v2_{ts}.xlsx"

    sheets = {
        "README": readme,
        "QC_Summary": qc_summary,
        "Study_Review": study,
        "Target_Control_Map_Review": target_control,
        "Problem_Rows": problems_actionable,
        "Metadata_Gaps": metadata_gaps,
        "Rowwise_Review": rowwise_df,
        "Sample_Map_Review": sample_map,
        "Technical_Inventory": technical,
    }

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        for name, df in sheets.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)
        autosize(writer, {k[:31]: v for k, v in sheets.items()})

    latest = OUTDIR / "LATEST_CHIP_CURATOR_REVIEW.txt"
    latest.write_text(str(out) + "\n")

    report = OUTDIR / f"chip_curator_review_v2_{ts}.summary.txt"
    report.write_text(
        "ChIP curator workbook V2\n"
        f"Generated: {datetime.now().isoformat(timespec='seconds')}\n"
        f"Workbook: {out}\n\n"
        f"Study_Review rows: {len(study)}\n"
        f"Target_Control_Map_Review rows: {len(target_control)}\n"
        f"Problem_Rows rows: {len(problems_actionable)}\n"
        f"Rowwise_Review rows: {len(rowwise_df)}\n"
        f"Metadata_Gaps rows: {len(metadata_gaps)}\n"
        f"Sample_Map_Review rows: {len(sample_map)}\n\n"
        "Improvements over V1:\n"
        "- Target_Control_Map_Review now includes target_confidence, target_review_flag, target_evidence.\n"
        "- Target_Control_Map_Review resolves control runs to control source rows/classes/roles when possible.\n"
        "- Problem_Rows now includes actionable curator-review issues with full row context, confidence, evidence, and control fields.\n"
        "- Metadata_Gaps separates unknown condition/stage rows from actionable problems.\n"
        "- Workbook avoids Excel structured tables to prevent Excel repair warnings.\n"
        "- Rowwise_Review includes source metadata and RNA-style curator_final_* fields.\n"
        "- Technical paths are moved to Technical_Inventory.\n"
    )

    print("Wrote:", out)
    print("Wrote:", latest)
    print("Wrote:", report)
    print()
    for name, df in sheets.items():
        print(f"{name}: {len(df)} rows, {len(df.columns)} cols")


if __name__ == "__main__":
    main()
