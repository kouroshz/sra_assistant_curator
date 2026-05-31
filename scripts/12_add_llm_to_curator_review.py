#!/usr/bin/env python3

from pathlib import Path
import argparse
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "outputs"


def clean(x):
    if pd.isna(x):
        return ""
    return str(x).strip()


def autosize(writer):
    wb = writer.book

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        for col_cells in ws.columns:
            max_len = 10
            letter = col_cells[0].column_letter
            for cell in col_cells:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            ws.column_dimensions[letter].width = min(max_len + 2, 80)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--pmid", required=True)
    args = parser.parse_args()

    pmid = clean(args.pmid)

    review_file = OUT / f"PMID_{pmid}_curator_review_view.xlsx"
    if not review_file.exists():
        raise FileNotFoundError(f"Missing review file: {review_file}. Run script 09 first.")

    summary_file = OUT / f"PMID_{pmid}_llm_summary.tsv"
    rows_file = OUT / f"PMID_{pmid}_llm_row_suggestions.tsv"
    groups_file = OUT / f"PMID_{pmid}_llm_group_suggestions.tsv"

    if not summary_file.exists():
        raise FileNotFoundError(f"Missing LLM summary: {summary_file}. Run script 11 first.")

    sheets = pd.read_excel(review_file, sheet_name=None, dtype=str)
    sheets = {k: v.fillna("") for k, v in sheets.items()}

    llm_summary = pd.read_csv(summary_file, sep="\t", dtype=str).fillna("")
    llm_rows = pd.read_csv(rows_file, sep="\t", dtype=str).fillna("") if rows_file.exists() else pd.DataFrame()
    llm_groups = pd.read_csv(groups_file, sep="\t", dtype=str).fillna("") if groups_file.exists() else pd.DataFrame()

    for sheet in ["Curator_View", "Needs_Review"]:
        if sheet in sheets and not llm_rows.empty and "Run" in sheets[sheet].columns:
            sheets[sheet] = sheets[sheet].merge(llm_rows, on="Run", how="left").fillna("")

    sheets["LLM_Summary"] = llm_summary
    sheets["LLM_Group_Suggestions"] = llm_groups
    sheets["LLM_Row_Suggestions"] = llm_rows

    out = OUT / f"PMID_{pmid}_curator_review_view_with_llm.xlsx"

    order = [
        "Curator_View",
        "Needs_Review",
        "Group_Summary",
        "Paper_Context",
        "LLM_Summary",
        "LLM_Group_Suggestions",
        "LLM_Row_Suggestions",
    ]

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        written = set()

        for s in order:
            if s in sheets:
                sheets[s].to_excel(writer, sheet_name=s[:31], index=False)
                written.add(s)

        for s, df in sheets.items():
            if s not in written:
                df.to_excel(writer, sheet_name=s[:31], index=False)

        autosize(writer)

    print(f"\n=== Added LLM output to curator review workbook for PMID {pmid} ===")
    print(f"Wrote: {out}")
    print(f"\nOpen with:\nopen {out}")


if __name__ == "__main__":
    main()
