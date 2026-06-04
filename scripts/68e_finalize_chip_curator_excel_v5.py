#!/usr/bin/env python3
"""
Final ChIP curator workbook polish.

Goals:
- Add RNA-style paper/study summaries.
- Add Paper_Summaries sheet.
- Add direct visible color fills, not only conditional formatting.
- Preserve V4 triage/problem aggregation.
- Avoid Excel structured tables to prevent repair warnings.
"""

from pathlib import Path
from datetime import datetime
import json
import re
import zipfile
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

OUTDIR = Path("outputs/06_CHIP_AI_ASSIST/21_curator_excel")
LATEST = OUTDIR / "LATEST_CHIP_CURATOR_REVIEW.txt"
ACTIVE = Path("outputs/06_CHIP_AI_ASSIST/14_chip_ai_inventory/chip_ai_active_validated_outputs.tsv")


COLORS = {
    "green": "D9EAD3",
    "yellow": "FFF2CC",
    "red": "F4CCCC",
    "orange": "FCE5CD",
    "blue": "DDEBF7",
    "gray": "E7E6E6",
    "header": "1F4E78",
    "white": "FFFFFF",
}


def clean(x):
    """Robust scalar/list/dict cleaner for workbook-facing text."""
    if x is None:
        return ""

    # AI JSON fields may be lists/dicts.
    if isinstance(x, list):
        vals = [clean(v) for v in x]
        vals = [v for v in vals if v]
        return " | ".join(vals)

    if isinstance(x, dict):
        vals = []
        for k, v in x.items():
            vv = clean(v)
            if vv:
                vals.append(f"{k}: {vv}")
        return " | ".join(vals)

    try:
        if pd.isna(x):
            return ""
    except Exception:
        pass

    s = str(x).strip()
    if s.lower() in {"nan", "none", "<na>"}:
        return ""
    return s


def norm(x):
    return clean(x).lower().replace("-", "_").replace(" ", "_")


def read_json(path):
    try:
        return json.loads(Path(path).read_text())
    except Exception:
        return {}


def first_nonblank(*vals):
    for v in vals:
        v = clean(v)
        if v:
            return v
    return ""


def short(x, n=1600):
    x = clean(x)
    return x[:n]


def get_nested(obj, *paths):
    for path in paths:
        cur = obj
        ok = True
        for key in path:
            if isinstance(cur, dict) and key in cur:
                cur = cur[key]
            else:
                ok = False
                break
        if ok:
            val = clean(cur) if not isinstance(cur, (dict, list)) else json.dumps(cur, ensure_ascii=False)
            if val:
                return val
    return ""


def summarize_list_field(obj, *paths):
    for path in paths:
        cur = obj
        ok = True
        for key in path:
            if isinstance(cur, dict) and key in cur:
                cur = cur[key]
            else:
                ok = False
                break
        if ok:
            if isinstance(cur, list):
                return " | ".join(clean(x) for x in cur if clean(x))
            if isinstance(cur, dict):
                return json.dumps(cur, ensure_ascii=False)
            return clean(cur)
    return ""


def fallback_summary(row):
    targets = clean(row.get("targets", ""))
    peak = clean(row.get("chip_peak_calling_ready", ""))
    n_rows = clean(row.get("n_rows", ""))
    if targets:
        return f"ChIP-like target-enrichment packet with {n_rows} rows for targets: {targets[:220]}; peak-calling readiness={peak}."
    return f"ChIP-like target-enrichment packet with {n_rows} rows; peak-calling readiness={peak}."


def extract_paper_summary(inv_row):
    obj = read_json(clean(inv_row.get("active_ai_json", "")))

    study_summary = obj.get("study_summary", {}) if isinstance(obj.get("study_summary", {}), dict) else {}
    paper_summary = obj.get("paper_summary", {}) if isinstance(obj.get("paper_summary", {}), dict) else {}
    ar = obj.get("analysis_readiness", {}) if isinstance(obj.get("analysis_readiness", {}), dict) else {}

    one_sentence = first_nonblank(
        get_nested(obj, ("one_sentence_summary",)),
        clean(study_summary.get("one_sentence_summary", "")),
        clean(study_summary.get("summary", "")),
        clean(paper_summary.get("one_sentence_summary", "")),
        clean(paper_summary.get("summary", "")),
        fallback_summary(inv_row),
    )

    study_goal = first_nonblank(
        get_nested(obj, ("study_goal",)),
        clean(study_summary.get("study_goal", "")),
        clean(study_summary.get("goal", "")),
        clean(paper_summary.get("study_goal", "")),
        clean(paper_summary.get("goal", "")),
    )

    organism_strain = first_nonblank(
        get_nested(obj, ("organism_strain",)),
        clean(study_summary.get("organism_strain", "")),
        clean(study_summary.get("organism_or_strain", "")),
        clean(study_summary.get("strain", "")),
        clean(paper_summary.get("organism_strain", "")),
    )

    comparisons = first_nonblank(
        get_nested(obj, ("main_comparisons_or_sample_axes",)),
        clean(study_summary.get("main_comparisons_or_sample_axes", "")),
        clean(study_summary.get("sample_axes", "")),
        clean(study_summary.get("main_comparisons", "")),
        clean(paper_summary.get("main_comparisons_or_sample_axes", "")),
    )

    evidence_locations = first_nonblank(
        summarize_list_field(obj, ("paper_evidence_locations",)),
        summarize_list_field(study_summary, ("paper_evidence_locations",)),
        summarize_list_field(paper_summary, ("paper_evidence_locations",)),
        summarize_list_field(obj, ("evidence_locations",)),
    )

    curator_warnings = first_nonblank(
        summarize_list_field(obj, ("curator_warnings",)),
        summarize_list_field(obj, ("global_warnings",)),
    )

    technical_warnings = first_nonblank(
        summarize_list_field(ar, ("main_blockers",)),
        summarize_list_field(obj, ("technical_warnings",)),
    )

    return {
        "packet_id": clean(inv_row.get("packet_id", "")),
        "pmid": clean(inv_row.get("pmid", "")),
        "bioproject": clean(inv_row.get("bioproject", "")),
        "targets": clean(inv_row.get("targets", "")),
        "chip_peak_calling_ready": clean(inv_row.get("chip_peak_calling_ready", "")),
        "one_sentence_summary": short(one_sentence),
        "study_goal": short(study_goal),
        "organism_strain": short(organism_strain),
        "main_comparisons_or_sample_axes": short(comparisons),
        "paper_evidence_locations": short(evidence_locations),
        "curator_warnings_clean": short(curator_warnings),
        "technical_warnings_clean": short(technical_warnings),
        "active_ai_json": clean(inv_row.get("active_ai_json", "")),
        "curator_summary_status": "",
        "curator_notes": "",
    }


def build_paper_summaries(active):
    rows = []
    for _, r in active.iterrows():
        rows.append(extract_paper_summary(r))
    return pd.DataFrame(rows)


def add_summary_columns_to_study(study, paper):
    study = study.copy()
    paper_small = paper[[
        "packet_id",
        "one_sentence_summary",
        "study_goal",
        "organism_strain",
        "main_comparisons_or_sample_axes",
        "paper_evidence_locations",
        "curator_warnings_clean",
        "technical_warnings_clean",
    ]].copy()

    # Remove old versions if rerunning.
    for c in paper_small.columns:
        if c != "packet_id" and c in study.columns:
            study = study.drop(columns=[c])

    study = study.merge(paper_small, on="packet_id", how="left")

    preferred = [
        "curator_priority",
        "curator_focus_summary",
        "packet_id",
        "pmid",
        "bioproject",
        "one_sentence_summary",
        "study_goal",
        "organism_strain",
        "main_comparisons_or_sample_axes",
        "paper_evidence_locations",
        "curator_warnings_clean",
        "technical_warnings_clean",
    ]
    preferred = [c for c in preferred if c in study.columns]
    rest = [c for c in study.columns if c not in preferred]
    return study[preferred + rest]


def color_for_value(value, col_name=""):
    v = norm(value)
    c = norm(col_name)

    if not v:
        return None

    # Confidence
    if "confidence" in c:
        if "low" in v:
            return COLORS["red"]
        if "medium" in v:
            return COLORS["yellow"]
        if "high" in v:
            return COLORS["green"]

    # Priority/severity/status/readiness/match
    if any(k in c for k in ["priority", "severity", "status", "ready", "readiness", "match", "flag", "warning"]):
        if any(k in v for k in ["high_review", "fail", "missing", "unresolved", "no"]):
            return COLORS["red"]
        if any(k in v for k in ["review", "partial", "curator_check", "ambiguous", "medium"]):
            return COLORS["yellow"] if "ambiguous" not in v else COLORS["orange"]
        if any(k in v for k in ["pass", "ok", "yes", "high"]):
            return COLORS["green"]

    # Problem flags/messages
    if "problem" in c:
        if any(k in v for k in ["missing", "unresolved", "low"]):
            return COLORS["red"]
        if any(k in v for k in ["unknown", "review", "partial", "curator_check", "ambiguous"]):
            return COLORS["yellow"]

    return None


def add_color_legend(sheets):
    legend = pd.DataFrame([
        {"color": "green", "meaning": "PASS / OK / yes / high confidence"},
        {"color": "yellow", "meaning": "REVIEW / partial / medium confidence / curator_check"},
        {"color": "red", "meaning": "HIGH_REVIEW / missing / unresolved / fail / no / low confidence"},
        {"color": "orange", "meaning": "ambiguous"},
        {"color": "blue", "meaning": "human-facing summary/info columns"},
    ])

    ordered = {}
    for name, df in sheets.items():
        ordered[name] = df
        if name == "README":
            ordered["Color_Legend"] = legend
    return ordered


def style_workbook(path):
    wb = load_workbook(path)

    header_fill = PatternFill("solid", fgColor=COLORS["header"])
    header_font = Font(bold=True, color=COLORS["white"])
    thin_gray = Border(bottom=Side(style="thin", color="D9D9D9"))

    freeze_map = {
        "README": "A2",
        "Color_Legend": "A2",
        "QC_Summary": "A2",
        "Curator_Triage": "D2",
        "Study_Review": "F2",
        "Paper_Summaries": "F2",
        "Target_Control_Map_Review": "G2",
        "Problem_Rows": "G2",
        "Problem_Details": "G2",
        "Metadata_Gaps": "G2",
        "Rowwise_Review": "G2",
        "Sample_Map_Review": "E2",
        "Technical_Inventory": "D2",
    }

    summary_cols = {
        "curator_focus_summary",
        "one_sentence_summary",
        "study_goal",
        "organism_strain",
        "main_comparisons_or_sample_axes",
        "paper_evidence_locations",
        "curator_warnings_clean",
        "technical_warnings_clean",
    }

    wrap_keywords = [
        "summary", "message", "hint", "evidence", "metadata", "warning",
        "notes", "comment", "json", "source_row_ids", "run_ids",
        "blockers", "description", "targets", "matched", "resolved",
        "comparisons", "goal", "locations"
    ]

    for ws in wb.worksheets:
        ws.freeze_panes = freeze_map.get(ws.title, "A2")

        if ws.max_row >= 2:
            ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_gray
        ws.row_dimensions[1].height = 30

        header_to_col = {str(ws.cell(row=1, column=i).value): i for i in range(1, ws.max_column + 1)}

        # Widths and direct fills
        for idx in range(1, ws.max_column + 1):
            col_name = str(ws.cell(row=1, column=idx).value)
            col_l = col_name.lower()

            vals = [col_name] + [str(ws.cell(row=r, column=idx).value or "") for r in range(2, min(ws.max_row, 200) + 1)]
            max_len = max(len(v) for v in vals)
            wrap = any(k in col_l for k in wrap_keywords)

            if col_name in summary_cols:
                width = 45
            elif wrap:
                width = min(max(max_len + 2, 22), 60)
            else:
                width = min(max(max_len + 2, 10), 32)

            ws.column_dimensions[get_column_letter(idx)].width = width

            for cell in ws[get_column_letter(idx)]:
                cell.alignment = Alignment(vertical="top", wrap_text=wrap)

            # Direct visible color fill for curator signal columns.
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=idx)
                fill = color_for_value(cell.value, col_name)
                if fill:
                    cell.fill = PatternFill("solid", fgColor=fill)

                # Light blue for high-level paper summary columns.
                if col_name in summary_cols and clean(cell.value):
                    cell.fill = PatternFill("solid", fgColor=COLORS["blue"])

        # Color legend direct fills
        if ws.title == "Color_Legend":
            for r in range(2, ws.max_row + 1):
                color_name = norm(ws.cell(row=r, column=1).value)
                if color_name in COLORS:
                    ws.cell(row=r, column=1).fill = PatternFill("solid", fgColor=COLORS[color_name])

        for ridx in range(2, ws.max_row + 1):
            ws.row_dimensions[ridx].height = 18

        # Keep conditional formatting too, but direct fills above make colors visible immediately.
        def col_letter(col_name):
            idx = header_to_col.get(col_name)
            return get_column_letter(idx) if idx else None

        def add_contains(col_name, text, fill):
            col = col_letter(col_name)
            if not col or ws.max_row < 2:
                return
            rng = f"{col}2:{col}{ws.max_row}"
            formula = f'IFERROR(ISNUMBER(SEARCH("{text}",{col}2)),FALSE)'
            ws.conditional_formatting.add(
                rng,
                FormulaRule(formula=[formula], fill=PatternFill("solid", fgColor=fill))
            )

        for c in ["suggestion_confidence", "confidence", "target_confidence", "resolved_control_confidence"]:
            add_contains(c, "low", COLORS["red"])
            add_contains(c, "medium", COLORS["yellow"])
            add_contains(c, "high", COLORS["green"])

        for c in [
            "priority", "curator_priority", "problem_severity", "severity",
            "review_flag", "target_review_flag", "resolved_control_review_flag",
            "warning_flags", "target_control_match_status",
            "validation_status", "chip_peak_calling_ready", "analysis_ready_status"
        ]:
            add_contains(c, "HIGH_REVIEW", COLORS["red"])
            add_contains(c, "FAIL", COLORS["red"])
            add_contains(c, "missing", COLORS["red"])
            add_contains(c, "unresolved", COLORS["red"])
            add_contains(c, "no", COLORS["red"])
            add_contains(c, "REVIEW", COLORS["yellow"])
            add_contains(c, "partial", COLORS["yellow"])
            add_contains(c, "curator_check", COLORS["yellow"])
            add_contains(c, "ambiguous", COLORS["orange"])
            add_contains(c, "PASS", COLORS["green"])
            add_contains(c, "OK", COLORS["green"])
            add_contains(c, "ok", COLORS["green"])
            add_contains(c, "yes", COLORS["green"])

    wb.save(path)


def validate_workbook(path):
    with zipfile.ZipFile(path) as z:
        tables = [n for n in z.namelist() if n.startswith("xl/tables/")]
        cf_count = 0
        for n in z.namelist():
            if n.startswith("xl/worksheets/") and n.endswith(".xml"):
                txt = z.read(n).decode("utf-8", errors="ignore")
                cf_count += txt.count("<conditionalFormatting")
    return len(tables), cf_count


def main():
    latest = Path(LATEST.read_text().strip())
    sheets = pd.read_excel(latest, sheet_name=None, dtype=str)
    sheets = {k: v.fillna("") for k, v in sheets.items()}

    active = pd.read_csv(ACTIVE, sep="\t", dtype=str).fillna("")
    paper = build_paper_summaries(active)

    if "Study_Review" in sheets:
        sheets["Study_Review"] = add_summary_columns_to_study(sheets["Study_Review"], paper)

    sheets["Paper_Summaries"] = paper

    preferred_order = [
        "README",
        "Color_Legend",
        "QC_Summary",
        "Curator_Triage",
        "Study_Review",
        "Paper_Summaries",
        "Target_Control_Map_Review",
        "Problem_Rows",
        "Problem_Details",
        "Metadata_Gaps",
        "Rowwise_Review",
        "Sample_Map_Review",
        "Technical_Inventory",
    ]

    sheets = add_color_legend(sheets)

    ordered = {}
    for name in preferred_order:
        if name in sheets:
            ordered[name] = sheets[name]
    for name, df in sheets.items():
        if name not in ordered:
            ordered[name] = df

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = OUTDIR / f"chip_curator_review_v5_{ts}.xlsx"

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        for name, df in ordered.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)

    style_workbook(out)
    table_count, cf_count = validate_workbook(out)

    LATEST.write_text(str(out) + "\n")

    summary = OUTDIR / f"{out.stem}.summary.txt"
    lines = [
        "ChIP curator workbook V5",
        f"Generated: {datetime.now().isoformat(timespec='seconds')}",
        f"Input: {latest}",
        f"Workbook: {out}",
        "",
    ]
    for name, df in ordered.items():
        lines.append(f"{name}: {len(df)} rows, {len(df.columns)} columns")
    lines.extend([
        "",
        f"Excel table parts: {table_count}",
        f"Conditional formatting blocks: {cf_count}",
        "",
        "V5 improvements:",
        "- Added Paper_Summaries sheet.",
        "- Added RNA-style paper summary columns to Study_Review.",
        "- Added direct visible color fills for confidence/status/review fields.",
        "- Kept conditional formatting as a backup, but colors are now visible immediately.",
        "- Added Color_Legend sheet.",
        "- No structured Excel tables, avoiding Excel repair warnings.",
    ])
    summary.write_text("\n".join(lines))

    print("Wrote:", out)
    print("Wrote:", summary)
    print()
    print("\n".join(lines))


if __name__ == "__main__":
    main()
