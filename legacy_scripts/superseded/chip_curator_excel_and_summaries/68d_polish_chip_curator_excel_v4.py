#!/usr/bin/env python3
"""
Final curator-facing polish for ChIP workbook.

Main goals:
- Make Problem_Rows match RNA style: one row per source row, aggregated flags/messages.
- Preserve original detailed issue rows in Problem_Details.
- Flag target-control rows where AI names a background class but no concrete control source row is resolved.
- Add Curator_Triage sheet to save curator time.
- Keep confidence/status highlighting and no Excel structured tables.
"""

from pathlib import Path
from datetime import datetime
from collections import Counter
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter

OUTDIR = Path("outputs/06_CHIP_AI_ASSIST/21_curator_excel")
LATEST = OUTDIR / "LATEST_CHIP_CURATOR_REVIEW.txt"


def clean(x):
    if pd.isna(x):
        return ""
    s = str(x).strip()
    if s.lower() in {"nan", "none", "<na>"}:
        return ""
    return s


def norm(x):
    return clean(x).lower().replace("-", "_").replace(" ", "_")


def join_unique(vals, max_len=2000):
    xs = []
    seen = set()
    for v in vals:
        v = clean(v)
        if v and v not in seen:
            xs.append(v)
            seen.add(v)
    return " | ".join(xs)[:max_len]


def max_severity(vals):
    vals = [clean(v) for v in vals]
    if "HIGH_REVIEW" in vals:
        return "HIGH_REVIEW"
    if "REVIEW" in vals:
        return "REVIEW"
    return vals[0] if vals else ""


def action_hint(flags):
    f = norm(flags)
    hints = []

    if "low_or_missing_confidence" in f:
        hints.append("Review AI call carefully; confirm/correct using source metadata and evidence.")
    if "missing_background" in f or "target_ip_missing_background" in f:
        hints.append("Resolve target-IP to input/IgG/background control or mark as not peak-calling-ready.")
    if "control_class_unresolved_to_run" in f:
        hints.append("AI identified a background class, but no concrete control SRR/source row was resolved; curator should verify.")
    if "review_flag_curator_check" in f:
        hints.append("Inspect because AI/deterministic repair explicitly requested curator check.")
    if "review_flag_ambiguous" in f:
        hints.append("Resolve ambiguity using paper/sample metadata.")
    if "packet_peak_calling_readiness" in f:
        hints.append("Review study-level ChIP readiness.")

    return " ".join(hints) or "Review source metadata, AI evidence, and curator notes before finalizing."


def aggregate_problem_rows(problem_details):
    if problem_details.empty:
        return problem_details.copy()

    df = problem_details.copy()
    df["_group_key"] = df["source_row_id"].where(df["source_row_id"].astype(str).str.len() > 0,
                                                "PACKET__" + df["packet_id"].astype(str) + "__" + df["problem_type"].astype(str))

    base_cols = [
        "packet_id", "pmid", "bioproject", "source_row_id", "Run", "BioSample",
        "target_clean", "suggested_target_or_antibody_or_tag",
        "sample_role_prelim", "suggested_sample_role",
        "stage", "strain", "condition",
        "suggested_comparator_or_background",
        "matched_background_run_ids_prelim",
        "assigned_control1", "assigned_control2",
        "suggestion_confidence", "review_flag",
        "suggestion_evidence", "raw_metadata_joined",
    ]
    base_cols = [c for c in base_cols if c in df.columns]

    rows = []
    for _, g in df.groupby("_group_key", sort=False):
        first = g.iloc[0].to_dict()

        row = {c: first.get(c, "") for c in base_cols}
        row["problem_severity"] = max_severity(g["severity"]) if "severity" in g.columns else ""
        row["problem_flags"] = join_unique(g["problem_type"]) if "problem_type" in g.columns else ""
        row["problem_messages"] = join_unique(g["problem_message"]) if "problem_message" in g.columns else ""
        row["curator_action_hint"] = action_hint(row["problem_flags"])
        row["n_problem_flags"] = len(set(clean(x) for x in g.get("problem_type", []) if clean(x)))
        row["curator_problem_status"] = ""
        row["curator_corrected_value"] = ""
        row["curator_notes"] = ""
        rows.append(row)

    preferred = [
        "problem_severity", "problem_flags", "problem_messages", "curator_action_hint",
        "n_problem_flags",
        "packet_id", "pmid", "bioproject", "source_row_id", "Run", "BioSample",
        "target_clean", "suggested_target_or_antibody_or_tag",
        "sample_role_prelim", "suggested_sample_role",
        "stage", "strain", "condition",
        "suggested_comparator_or_background",
        "matched_background_run_ids_prelim",
        "assigned_control1", "assigned_control2",
        "suggestion_confidence", "review_flag",
        "suggestion_evidence", "raw_metadata_joined",
        "curator_problem_status", "curator_corrected_value", "curator_notes",
    ]
    out = pd.DataFrame(rows)
    preferred = [c for c in preferred if c in out.columns]
    rest = [c for c in out.columns if c not in preferred]
    return out[preferred + rest]


def add_unresolved_control_problems(target_control, problem_details):
    if target_control.empty:
        return problem_details

    add_rows = []
    for _, r in target_control.iterrows():
        status = clean(r.get("target_control_match_status", ""))
        resolved = clean(r.get("resolved_control_source_row_ids", ""))
        ai_bg = clean(r.get("ai_background_or_comparator", ""))
        prelim = clean(r.get("prelim_matched_background_run_ids", ""))

        # Suspicious case: AI has a comparator class, but we cannot point to concrete control rows.
        if (
            status == "ok"
            and not resolved
            and ai_bg
            and norm(ai_bg) not in {"unknown", "not_applicable", "na", "n_a", "none"}
        ):
            add_rows.append({
                "problem_type": "control_class_unresolved_to_run",
                "severity": "REVIEW",
                "problem_message": (
                    f"control_class_unresolved_to_run; Run={clean(r.get('target_Run'))}; "
                    f"target={clean(r.get('suggested_target_or_antibody_or_tag'))}; "
                    f"AI_bg={ai_bg}; prelim_bg={prelim or 'blank'}"
                ),
                "curator_action_hint": "AI identified a background/control class, but no concrete control SRR/source row was resolved.",
                "packet_id": clean(r.get("packet_id")),
                "pmid": clean(r.get("pmid")),
                "bioproject": clean(r.get("bioproject")),
                "source_row_id": clean(r.get("target_source_row_id")),
                "Run": clean(r.get("target_Run")),
                "BioSample": clean(r.get("target_BioSample")),
                "target_clean": clean(r.get("target_clean")),
                "suggested_target_or_antibody_or_tag": clean(r.get("suggested_target_or_antibody_or_tag")),
                "sample_role_prelim": "target_ip",
                "suggested_sample_role": "target_ip",
                "stage": clean(r.get("target_stage")),
                "strain": clean(r.get("target_strain")),
                "condition": clean(r.get("target_condition")),
                "suggested_comparator_or_background": ai_bg,
                "matched_background_run_ids_prelim": prelim,
                "assigned_control1": clean(r.get("assigned_control1")),
                "assigned_control2": clean(r.get("assigned_control2")),
                "suggestion_confidence": clean(r.get("target_confidence")),
                "review_flag": clean(r.get("target_review_flag")),
                "suggestion_evidence": clean(r.get("target_evidence")),
                "raw_metadata_joined": "",
                "curator_problem_status": "",
                "curator_corrected_value": "",
                "curator_notes": "",
            })

            target_control.loc[r.name, "target_control_match_status"] = "control_class_unresolved_to_run"

    if add_rows:
        problem_details = pd.concat([problem_details, pd.DataFrame(add_rows)], ignore_index=True)

    return problem_details


def build_triage(study, problem_rows, metadata_gaps, target_control):
    rows = []

    if not study.empty:
        for _, r in study.iterrows():
            priority = clean(r.get("curator_priority", ""))
            if priority in {"HIGH_REVIEW", "REVIEW"}:
                rows.append({
                    "triage_type": "study_priority",
                    "priority": priority,
                    "packet_id": clean(r.get("packet_id")),
                    "pmid": clean(r.get("pmid")),
                    "bioproject": clean(r.get("bioproject")),
                    "what_to_review": clean(r.get("curator_focus_summary")),
                    "suggested_sheet": "Study_Review",
                })

    if not problem_rows.empty:
        high = problem_rows[problem_rows["problem_severity"].astype(str).str.contains("HIGH", na=False)]
        rows.append({
            "triage_type": "actionable_rows",
            "priority": "HIGH_REVIEW",
            "packet_id": "",
            "pmid": "",
            "bioproject": "",
            "what_to_review": f"{len(high)} high-review row-level issues in Problem_Rows; {len(problem_rows)} total unique actionable rows.",
            "suggested_sheet": "Problem_Rows",
        })

    if not target_control.empty and "target_control_match_status" in target_control.columns:
        bad = target_control[target_control["target_control_match_status"] != "ok"]
        rows.append({
            "triage_type": "target_control",
            "priority": "REVIEW" if len(bad) else "OK",
            "packet_id": "",
            "pmid": "",
            "bioproject": "",
            "what_to_review": f"{len(bad)} target-control rows need review out of {len(target_control)}.",
            "suggested_sheet": "Target_Control_Map_Review",
        })

    rows.append({
        "triage_type": "metadata_gaps",
        "priority": "LOWER_PRIORITY",
        "packet_id": "",
        "pmid": "",
        "bioproject": "",
        "what_to_review": f"{len(metadata_gaps)} metadata gap rows, mostly unknown condition/stage. Review if needed for downstream harmonization.",
        "suggested_sheet": "Metadata_Gaps",
    })

    return pd.DataFrame(rows)


def style_workbook(path):
    wb = load_workbook(path)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    thin_gray = Border(bottom=Side(style="thin", color="D9D9D9"))

    freeze_map = {
        "README": "A2",
        "QC_Summary": "A2",
        "Curator_Triage": "D2",
        "Study_Review": "D2",
        "Target_Control_Map_Review": "G2",
        "Problem_Rows": "G2",
        "Problem_Details": "G2",
        "Metadata_Gaps": "G2",
        "Rowwise_Review": "G2",
        "Sample_Map_Review": "E2",
        "Technical_Inventory": "D2",
    }

    wrap_keywords = [
        "summary", "message", "hint", "evidence", "metadata", "warning",
        "notes", "comment", "json", "source_row_ids", "run_ids",
        "blockers", "description", "targets", "matched", "resolved"
    ]

    for ws in wb.worksheets:
        ws.freeze_panes = freeze_map.get(ws.title, "A2")
        if ws.max_row >= 2:
            ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_gray
        ws.row_dimensions[1].height = 30

        header_to_col = {str(ws.cell(row=1, column=i).value): i for i in range(1, ws.max_column + 1)}

        for idx in range(1, ws.max_column + 1):
            col_name = str(ws.cell(row=1, column=idx).value)
            col_l = col_name.lower()
            vals = [col_name] + [str(ws.cell(row=r, column=idx).value or "") for r in range(2, min(ws.max_row, 200) + 1)]
            max_len = max(len(v) for v in vals)
            wrap = any(k in col_l for k in wrap_keywords)
            width = min(max(max_len + 2, 22 if wrap else 10), 60 if wrap else 32)
            ws.column_dimensions[get_column_letter(idx)].width = width
            for cell in ws[get_column_letter(idx)]:
                cell.alignment = Alignment(vertical="top", wrap_text=wrap)

        for ridx in range(2, ws.max_row + 1):
            ws.row_dimensions[ridx].height = 18

        def col_letter(col_name):
            idx = header_to_col.get(col_name)
            return get_column_letter(idx) if idx else None

        def add_contains(col_name, text, fill):
            col = col_letter(col_name)
            if not col or ws.max_row < 2:
                return
            rng = f"{col}2:{col}{ws.max_row}"
            formula = f'IFERROR(ISNUMBER(SEARCH("{text}",{col}2)),FALSE)'
            ws.conditional_formatting.add(
                rng,
                FormulaRule(formula=[formula], fill=PatternFill("solid", fgColor=fill))
            )

        for c in [
            "suggestion_confidence", "confidence", "target_confidence",
            "resolved_control_confidence"
        ]:
            add_contains(c, "low", "F4CCCC")
            add_contains(c, "medium", "FFF2CC")
            add_contains(c, "high", "D9EAD3")

        for c in [
            "priority", "curator_priority", "problem_severity", "severity",
            "review_flag", "target_review_flag", "resolved_control_review_flag",
            "warning_flags", "target_control_match_status",
            "validation_status", "chip_peak_calling_ready", "analysis_ready_status"
        ]:
            add_contains(c, "HIGH_REVIEW", "F4CCCC")
            add_contains(c, "FAIL", "F4CCCC")
            add_contains(c, "missing", "F4CCCC")
            add_contains(c, "unresolved", "FFF2CC")
            add_contains(c, "low", "F4CCCC")
            add_contains(c, "no", "F4CCCC")
            add_contains(c, "REVIEW", "FFF2CC")
            add_contains(c, "partial", "FFF2CC")
            add_contains(c, "curator_check", "FFF2CC")
            add_contains(c, "ambiguous", "FCE5CD")
            add_contains(c, "PASS", "D9EAD3")
            add_contains(c, "OK", "D9EAD3")
            add_contains(c, "ok", "D9EAD3")
            add_contains(c, "yes", "D9EAD3")

    wb.save(path)


def main():
    latest = Path(LATEST.read_text().strip())
    sheets = pd.read_excel(latest, sheet_name=None, dtype=str)
    sheets = {k: v.fillna("") for k, v in sheets.items()}

    problem_details = sheets.get("Problem_Rows", pd.DataFrame()).copy()
    target_control = sheets.get("Target_Control_Map_Review", pd.DataFrame()).copy()

    problem_details = add_unresolved_control_problems(target_control, problem_details)
    problem_rows = aggregate_problem_rows(problem_details)

    metadata_details = sheets.get("Metadata_Gaps", pd.DataFrame()).copy()
    metadata_gaps = aggregate_problem_rows(metadata_details)

    study = sheets.get("Study_Review", pd.DataFrame()).copy()
    triage = build_triage(study, problem_rows, metadata_gaps, target_control)

    # Update QC summary with final counts.
    qc = sheets.get("QC_Summary", pd.DataFrame()).copy()
    extra_qc = pd.DataFrame([
        {"section": "Final V4", "metric": "Unique actionable Problem_Rows", "value": len(problem_rows)},
        {"section": "Final V4", "metric": "Problem_Details rows", "value": len(problem_details)},
        {"section": "Final V4", "metric": "Unique Metadata_Gaps rows", "value": len(metadata_gaps)},
        {"section": "Final V4", "metric": "Target-control rows needing review", "value": int((target_control.get("target_control_match_status", "") != "ok").sum()) if not target_control.empty else 0},
    ])
    qc = pd.concat([qc, extra_qc], ignore_index=True)

    ordered = {
        "README": sheets.get("README", pd.DataFrame()),
        "QC_Summary": qc,
        "Curator_Triage": triage,
        "Study_Review": study,
        "Target_Control_Map_Review": target_control,
        "Problem_Rows": problem_rows,
        "Problem_Details": problem_details,
        "Metadata_Gaps": metadata_gaps,
        "Rowwise_Review": sheets.get("Rowwise_Review", pd.DataFrame()),
        "Sample_Map_Review": sheets.get("Sample_Map_Review", pd.DataFrame()),
        "Technical_Inventory": sheets.get("Technical_Inventory", pd.DataFrame()),
    }

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = OUTDIR / f"chip_curator_review_v4_{ts}.xlsx"

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        for name, df in ordered.items():
            df.to_excel(writer, sheet_name=name[:31], index=False)

    style_workbook(out)
    LATEST.write_text(str(out) + "\n")

    summary = OUTDIR / f"{out.stem}.summary.txt"
    lines = [
        "ChIP curator workbook V4",
        f"Generated: {datetime.now().isoformat(timespec='seconds')}",
        f"Input: {latest}",
        f"Workbook: {out}",
        "",
    ]
    for name, df in ordered.items():
        lines.append(f"{name}: {len(df)} rows, {len(df.columns)} columns")
    lines.extend([
        "",
        "V4 improvements:",
        "- Problem_Rows is aggregated RNA-style: one row per source row/packet issue.",
        "- Original unaggregated issue rows preserved in Problem_Details.",
        "- Target-control rows with AI comparator class but no resolved control run are flagged.",
        "- Curator_Triage added as start-here sheet.",
    ])
    summary.write_text("\n".join(lines))

    print("Wrote:", out)
    print("Wrote:", summary)
    print()
    print("\n".join(lines))


if __name__ == "__main__":
    main()
